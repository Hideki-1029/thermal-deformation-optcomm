"""
orbit_catalog.xlsx column layout and Excel styling.

Human-edited (white)
--------------------
- Identity / Keplerian / notes
- Attitude intent:
  - ``sun_face``: body face → Sun
  - ``constraint_target``: ``velocity`` | ``nadir`` (what the second axis tracks)
  - ``constraint_face``: body face → that second target

Script-filled (gray) — do not hand-edit
---------------------------------------
- Effective faces: ``eff_sun_face``, ``eff_velocity_face``, ``eff_nadir_face``, sources
- TD dump: ``orient_type``, ``constraint_type``, pointing / constraint axes, Rot*, notes

``constraint_target`` vs ``constraint_type``
-------------------------------------------
- ``constraint_target``: human label ``velocity`` | ``nadir`` (input; Excel→TD later)
- ``constraint_type``: raw TD enum ``VELOCITY`` | ``PLANET`` (dump only; same meaning)
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Attitude intent (edit these)
INPUT_ATTITUDE_COLS = (
    "sun_face",
    "constraint_target",
    "constraint_face",
)

# What you care about as results (script fills)
RESULT_FACE_COLS = (
    "eff_sun_face",
    "eff_velocity_face",
    "eff_nadir_face",
    "eff_velocity_source",
    "eff_nadir_source",
)

# TD GUI / OpenTD dump (script fills; not an input surface)
TD_INTERNAL_COLS = (
    "orient_type",
    "constraint_type",
    "pointing_axis",
    "constraint_axis",
    "rot1_axis_code",
    "rot1_deg",
    "rot2_axis_code",
    "rot2_deg",
    "rot3_axis_code",
    "rot3_deg",
    "notes_attitude",
)

# Orbit-error / PAT wiring (script fills; see run_orbit_error_stt_frame.py)
ORBIT_ERROR_COLS = (
    "pat_orbit_error_frame",
    "orbit_error_stt_frame",
    "orbit_error_partner_mode",
    "orbit_error_stt_status",
    "orbit_error_stt_notes",
)

DERIVED_COLS = RESULT_FACE_COLS + TD_INTERNAL_COLS + ORBIT_ERROR_COLS

# Preferred left-to-right order for the active sheet
PREFERRED_COLUMN_ORDER = (
    "no",
    "td_orbit_name",
    "sun_face",
    "constraint_target",
    "constraint_face",
    "orbit_type",
    "inclination_deg",
    "raan_deg",
    "epoch_utc",
    "min_alt_km",
    "max_alt_km",
    "solar_w_per_mm2",
    "albedo",
    "ir_w_per_mm2",
    "illumination_condition",
    "orbit_period_s",
    "beta_angle_deg",
    "used_by_case_count",
    "notes",
    *RESULT_FACE_COLS,
    *TD_INTERNAL_COLS,
    *ORBIT_ERROR_COLS,
)

_DERIVED_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_DERIVED_HEADER_FONT = Font(bold=True, color="595959")
_INPUT_HEADER_FONT = Font(bold=True)


def reorder_orbit_catalog_columns(df: pd.DataFrame) -> pd.DataFrame:
    preferred = [c for c in PREFERRED_COLUMN_ORDER if c in df.columns]
    rest = [c for c in df.columns if c not in preferred]
    return df.loc[:, preferred + rest].copy()


def style_derived_columns(ws: Worksheet, *, header_row: int = 1) -> None:
    """Gray-fill script-filled columns; leave input columns white."""
    headers = [cell.value for cell in ws[header_row]]
    derived = set(DERIVED_COLS)
    max_row = ws.max_row or header_row
    for col_idx, name in enumerate(headers, start=1):
        if name not in derived:
            cell = ws.cell(header_row, col_idx)
            cell.font = _INPUT_HEADER_FONT
            continue
        letter = get_column_letter(col_idx)
        for row in range(header_row, max_row + 1):
            cell = ws.cell(row, col_idx)
            cell.fill = _DERIVED_FILL
            if row == header_row:
                cell.font = _DERIVED_HEADER_FONT


def write_orbit_catalog_xlsx(
    path: Path,
    catalog: pd.DataFrame,
    *,
    sheet_name: str = "orbit_catalog",
    archive: pd.DataFrame | None = None,
) -> None:
    ordered = reorder_orbit_catalog_columns(catalog)
    path = Path(path)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        ordered.to_excel(writer, sheet_name=sheet_name, index=False)
        if archive is not None:
            archive.to_excel(writer, sheet_name="orbit_catalog_archive", index=False)
        ws = writer.sheets[sheet_name]
        style_derived_columns(ws)


def load_archive_sheet(path: Path) -> pd.DataFrame | None:
    try:
        return pd.read_excel(path, sheet_name="orbit_catalog_archive")
    except Exception:
        return None
